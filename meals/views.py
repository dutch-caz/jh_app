import openpyxl
from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseRedirect
from .models import attendee
from .models import meal
from .models import attendance
from .models import enableSignUp
from .models import announcement
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

from django.utils import timezone

def generate_attendance(request):
    # Create a new workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Crew Meal Attendance'
    
    # Write column headers
    worksheet['A1'] = 'Name'
    worksheet['B1'] = 'Type'
    
    # Get meal objects sorted by start date
    meals = meal.objects.order_by('start_time__date')
    
    # Map each date to its starting column
    date_to_column_map = {}
    base_column = 3  # Start columns for meal headers at column 3
    
    # Define the red fill style for cells
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    
    # Iterate through meals and map dates to columns
    for meal_obj in meals:
        date_str = meal_obj.start_time.strftime('%m/%d')
        if date_str not in date_to_column_map:
            # Calculate the starting column for this date
            start_column = base_column + len(date_to_column_map) * 3
            date_to_column_map[date_str] = start_column
            
            # Merge cells for date header
            worksheet.merge_cells(f'{get_column_letter(start_column)}1:{get_column_letter(start_column + 2)}1')
            worksheet[f'{get_column_letter(start_column)}1'] = date_str
            
            # Set the BLD headers
            worksheet[f'{get_column_letter(start_column)}2'] = 'B'
            worksheet[f'{get_column_letter(start_column + 1)}2'] = 'L'
            worksheet[f'{get_column_letter(start_column + 2)}2'] = 'D'
    
    # Get all attendees sorted by name
    attendees = attendee.objects.order_by('name')
    
    # Write attendee names and attendance data
    row = 3
    column = 3
    
    for attendee_obj in attendees:
        # Write attendee name and type
        worksheet.cell(row=row, column=1, value=attendee_obj.name)
        worksheet.cell(row=row, column=2, value=attendee_obj.person_type)
        
        # Check each meal
        for meal_obj in meals:
            attendance_obj = attendance.objects.filter(attendee=attendee_obj, meal=meal_obj).first()
            if attendance_obj:
                worksheet.cell(row=row, column=column, value=1.0 + attendance_obj.attendee.half_plates)
                column += 1
                if attendance_obj.present == False:
                    worksheet.cell(row=row, column=column - 1).fill = red_fill
            else:
                worksheet.cell(row=row, column=column, value=0.0)
                column += 1
        # Move to the next row for the next attendee
        column = 3
        row += 1
    
    # Auto fit column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.coordinate in worksheet.merged_cells:
                continue
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        worksheet.column_dimensions[column_letter].width = adjusted_width
        
    # Set column widths for columns after C
    # Get the maximum column index
    max_column_index = max([column[0].column for column in worksheet.columns])

    # Set the width of columns starting from the 4th column
    for column_index in range(3, max_column_index + 1):
        column_letter = openpyxl.utils.get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = 5

    # Turn off rounding for every cell
    for row in worksheet.iter_rows(min_row=1, min_col=1):
        for cell in row:
            cell.number_format = '0.00'

    # Turn on text wrapping for every cell
    for row in worksheet.iter_rows(min_row=1, min_col=1):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    # Center align every cell
    for row in worksheet.iter_rows(min_row=1, min_col=1):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    
    # Set the content type and filename for the response
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response = HttpResponse(content_type=content_type)
    response['Content-Disposition'] = 'attachment; filename=JH_Attendance.xlsx'
    
    # Save the workbook to the response
    workbook.save(response)
    
    return response
    

def generate_spreadsheet(request):
    # Create a new workbook and worksheet
    workbook = openpyxl.load_workbook('C:/Users/dutch/Desktop/Programming/JH Ranch/jhango/jh_app/meals/templates/JH_Meal_Template.xlsx')
    crew_plates_worksheet = workbook['Crew_Plates']

    # Write column headers
    crew_plates_worksheet['A1'] = 'Name'
    crew_plates_worksheet['B1'] = 'Type'
    crew_plates_worksheet['C1'] = 'Dietary Restrictions'

    # Get meal objects sorted by start date
    meals = meal.objects.order_by('start_time__date')

    # Write headers for each meal
    dates = set()
    for meal_obj in meals:
        date_str = meal_obj.start_time.strftime('%m/%d')
        if date_str not in dates:
            column = len(dates) * 3 + 4
            column_letter = get_column_letter(column)
            crew_plates_worksheet.merge_cells(f'{column_letter}1:{get_column_letter(column+2)}1')
            crew_plates_worksheet[f'{column_letter}1'] = date_str

            dates.add(date_str)
        
    # Set column widths for merged cells
    merged_columns = len(dates) * 3
    column_width = (crew_plates_worksheet.column_dimensions['A'].width + 2) * 1.1
    for i in range(merged_columns):
        column_letter = get_column_letter(i + 4)
        crew_plates_worksheet.column_dimensions[column_letter].width = column_width / 3
        
    # Create the B, L, and D columns under the merged cells
    for i in range(merged_columns // 3):
        column_letter = get_column_letter(i * 3 + 4)
        crew_plates_worksheet[column_letter + '2'] = 'B'
        crew_plates_worksheet[get_column_letter(i * 3 + 5) + '2'] = 'L'
        crew_plates_worksheet[get_column_letter(i * 3 + 6) + '2'] = 'D'
    
    # Get all attendees sorted by name
    attendees = attendee.objects.order_by('name')

    # Write attendee names alphabetically starting from the 3rd row
    row = 3
    dates = list(dates)  # Convert set to list

    # Logic to each users meal count
    for attendee_obj in attendees:
        # Name and Type columns
        crew_plates_worksheet.cell(row=row, column=1, value=attendee_obj.name)
        crew_plates_worksheet.cell(row=row, column=2, value=attendee_obj.person_type)
        # GF/DF logic
        crew_plates_worksheet.cell(row=row, column=3, value=1 if attendee_obj.gluten_free or attendee_obj.dairy_free or attendee_obj.nut_allergy or attendee_obj.vegan or attendee_obj.vegetarian else 0)
        
        # Check attendance for all meals
        attendance_values = []
        for meal_obj in meals:
            attendance_obj = attendance.objects.filter(attendee=attendee_obj, meal=meal_obj).first()
            if attendance_obj:
                attendance_values.append(1.0 + attendance_obj.attendee.half_plates)
            else:
                attendance_values.append(0.0)
        
        # Write attendance values for each meal date
        for i, attendance_value in enumerate(attendance_values):
            column = i + 4  # Assuming column offset starts from 4
            column_letter = get_column_letter(column)
            crew_plates_worksheet[f'{column_letter}{row}'] = attendance_value
            
        # Write notes for each user
        crew_plates_worksheet.cell(row=row, column=column + 1, value=attendee_obj.notes)
        
        row += 1 # Move to the next user row
        
    # Add a row for Normal plates
    crew_plates_worksheet.cell(row=row, column=1, value='Normal plates')
    crew_plates_worksheet.cell(row=row+1, column=1, value='GF/DF plates')

    # Generate formulas under each B, L, or D column
    for i in range(merged_columns):
        column = i + 4
        column_letter = get_column_letter(column)
        
        # Formula for Normal plates
        normal_formula = f'=SUMIF(C3:C{row-1}, 0.0, {column_letter}3:{column_letter}{row-1})'
        crew_plates_worksheet[f'{column_letter}{row}'] = normal_formula
        
        # Formula for GF/DF plates
        gf_df_formula = f'=SUMIF(C3:C{row-1}, 1.0, {column_letter}3:{column_letter}{row-1})'
        crew_plates_worksheet[f'{column_letter}{row+1}'] = gf_df_formula
    
    #Styling
    # Auto fit column widths
    for column in crew_plates_worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        crew_plates_worksheet.column_dimensions[column_letter].width = adjusted_width
        
    # Set column widths for columns after C
    # Get the maximum column index
    max_column_index = max([column[0].column for column in crew_plates_worksheet.columns])

    # Set the width of columns starting from the 4th column
    for column_index in range(4, max_column_index + 1):
        column_letter = openpyxl.utils.get_column_letter(column_index)
        crew_plates_worksheet.column_dimensions[column_letter].width = 5

    # Turn off rounding for every cell
    for row in crew_plates_worksheet.iter_rows(min_row=1, min_col=1):
        for cell in row:
            cell.number_format = '0.0'

    # Turn on text wrapping for every cell
    for row in crew_plates_worksheet.iter_rows(min_row=1, min_col=1):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    # Center align every cell
    for row in crew_plates_worksheet.iter_rows(min_row=1, min_col=1):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
    meal_data_worksheet = workbook['Meal_Data']
    
    


    # Set the content type for the response
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    # Create the response
    response = HttpResponse(content_type=content_type)
    response['Content-Disposition'] = 'attachment; filename=JH_Meals.xlsx'

    # Write the workbook to the response
    workbook.save(response)

    return response

def home(request):
    attendees = attendee.objects.order_by('name')
    announcements = announcement.objects.all()
    return render(request, 'meals/home.html', {'attendees' : attendees, 'announcements' : announcements})

def signup(request):
    meals = meal.objects.order_by('start_time__date')
    users = attendee.objects.all()
    enable = enableSignUp.objects.get(pk=2)
    
    if(request.method == "POST"):
        name = request.POST.get('attendee_name')
        user = attendee.objects.get(name=name)
        print(f'User: {user}')
        
        for meal_name in request.POST.getlist('meals'):
            meal_obj = meal.objects.get(name=meal_name)
            
            print(f'Meal: {meal_obj}')
            
            if request.POST.get(meal_name, True):
                new_attendance, created = attendance.objects.get_or_create(
                    attendee = user,
                    meal = meal_obj,
                    present = False
                )
                if created:
                    new_attendance.save()
                
        return redirect('home')
    
    return render(request, 'meals/signup.html', {'meals': meals, 'enable' : enable, 'users' : users})

def getcoords(request):
    return render(request, 'meals/getcoords.html', {})

def check_in(request):
    meals = meal.objects.order_by('start_time__date')
    users = attendee.objects.all()
    attendances = attendance.objects.all()
    current_time = timezone.now()

    if request.method == "POST":
        name = request.POST.get('user')
        meal_name = request.POST.get('meal')
        get_latitude = request.POST.get('latitude')
        get_longitude = request.POST.get('longitude')
        user = attendee.objects.get(name=name)
        meal_obj = meal.objects.get(name=meal_name)
        attendance_obj = attendance.objects.get(attendee=user, meal=meal_obj)
        if (float(get_latitude) - float(attendance_obj.meal.latitude))**2 + (float(get_longitude) - float(attendance_obj.meal.longitude))**2 < (0.0025)**2:
            attendance_obj.present = True
            attendance_obj.save()
            return redirect('success')
        else:
            return redirect('too_far')
    
    return render(request, 'meals/check_in.html', {'meals': meals, 'users' : users, 'current_time' : current_time,
                                                   'attendances' : attendances})
    
    
def success(request):
    return render(request, 'meals/success.html', {})

def too_far(request):
    return render(request, 'meals/too_far.html', {})
    